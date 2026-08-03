"""Score real Smith-Waterman masks against Excel subword bounding boxes.

The evaluator treats the order-preserving shared subwords between the two
annotated lines as ground-truth positives. A box is a predicted positive when
it falls inside the dense Smith-Waterman mask. This provides object-level
TP/FP/FN/TN, precision, recall, F1, specificity and accuracy, together with
interval and pixel-mask localization metrics.
"""
from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
import math
import os
from pathlib import Path
import re
import unicodedata

import numpy as np
from PIL import Image

from Evaluation._eval_utils import patch_range_to_pixels


_ARABIC_DIACRITICS = re.compile("[\u0610-\u061a\u064b-\u065f\u0670\u06d6-\u06ed]")
_LINE_NUMBER = re.compile(r"(?:line[_\-\s]*)?0*(\d+)", re.IGNORECASE)
_NUMBER = re.compile(r"[-+]?\d+(?:\.\d+)?")

_HEADER_ALIASES = {
    "text": {
        "subword", "sub_word", "connected_subword", "word", "token", "text",
        "label", "transcription", "content", "arabic", "string", "نص", "كلمة",
        "مقطع", "الكلمة",
    },
    "line": {
        "line", "line_idx", "line_index", "line_number", "line_no", "row",
        "row_idx", "row_index", "سطر", "رقم_السطر",
    },
    "image": {
        "image", "image_name", "filename", "file_name", "line_image",
        "image_path", "path", "اسم_الصورة",
    },
    "bbox": {
        "bbox", "bounding_box", "boundingbox", "box", "coordinates", "coords",
        "rectangle", "المربع", "الاحداثيات",
    },
    "x0": {"x0", "xmin", "x_min", "left", "start_x", "x_start", "min_x", "bbox_x0", "box_x0"},
    "y0": {"y0", "ymin", "y_min", "top", "start_y", "y_start", "min_y", "bbox_y0", "box_y0"},
    "x1": {"x1", "xmax", "x_max", "right", "end_x", "x_end", "max_x", "bbox_x1", "box_x1"},
    "y1": {"y1", "ymax", "y_max", "bottom", "end_y", "y_end", "max_y", "bbox_y1", "box_y1"},
    "x": {"x", "bbox_x", "box_x"},
    "y": {"y", "bbox_y", "box_y"},
    "width": {"w", "width", "bbox_width", "box_width"},
    "height": {"h", "height", "bbox_height", "box_height"},
}


@dataclass(frozen=True)
class SubwordBox:
    text: str
    x0: float
    y0: float
    x1: float
    y1: float
    source_row: int = -1

    @property
    def width(self) -> float:
        return max(0.0, self.x1 - self.x0)

    @property
    def height(self) -> float:
        return max(0.0, self.y1 - self.y0)

    @property
    def center_x(self) -> float:
        return 0.5 * (self.x0 + self.x1)


@dataclass(frozen=True)
class BoxAnnotations:
    boxes: tuple[SubwordBox, ...]
    workbook: str = ""
    sheet: str = ""
    status: str = "missing"
    error: str = ""


def _flag(name: str, default: bool) -> bool:
    value = os.environ.get(name)
    if value is None:
        return bool(default)
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _number(name: str, default: float) -> float:
    try:
        return float(os.environ.get(name, default))
    except (TypeError, ValueError):
        return float(default)


def _normalise_header(value) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-/\\]+", "_", text)
    return re.sub(r"[^\w]+", "", text, flags=re.UNICODE)


def _canonical_header(value) -> str | None:
    normalised = _normalise_header(value)
    for canonical, aliases in _HEADER_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _normalise_text(value) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = _ARABIC_DIACRITICS.sub("", text.replace("ـ", ""))
    return "".join(text.split())


def _as_float(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _line_index(image_path: Path) -> int | None:
    match = _LINE_NUMBER.search(image_path.stem)
    return int(match.group(1)) if match else None


def _side_root(image_path: Path) -> Path:
    for parent in image_path.parents:
        if parent.name in {"A", "B"}:
            return parent
    return image_path.parent.parent


def _candidate_score(path: Path, image_path: Path) -> tuple[int, int, str]:
    rendered = str(path).lower()
    score = 100 if image_path.stem.lower() in path.stem.lower() else 0
    for token, weight in (
        ("subword", 50), ("bbox", 45), ("bounding", 40), ("box", 35),
        ("annotation", 30), ("label", 20),
    ):
        if token in rendered:
            score += weight
    return (-score, len(path.parts), str(path))


@lru_cache(maxsize=4096)
def discover_workbook(image_value: str) -> Path | None:
    image_path = Path(image_value)
    side_root = _side_root(image_path)
    explicit_root = os.environ.get("REAL_BOX_ANNOTATIONS_ROOT", "").strip()
    roots = [Path(explicit_root).expanduser()] if explicit_root else []
    roots.append(side_root)

    names = (
        f"{image_path.stem}.xlsx", f"{image_path.stem}.xlsm", f"{image_path.stem}.xls",
        "subword_boxes.xlsx", "bounding_boxes.xlsx", "boxes.xlsx",
        "annotations.xlsx", "labels.xlsx",
    )
    candidates: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for directory in (
            root, root / "annotations", root / "Annotations", root / "bounding_boxes",
            root / "boxes", root / "excel", root / "Excel", root / "labels",
        ):
            for name in names:
                candidate = directory / name
                if candidate.is_file():
                    candidates.append(candidate)
        if not candidates:
            for pattern in ("*.xlsx", "*.xlsm", "*.xls"):
                candidates.extend(root.rglob(pattern))

    unique = sorted(
        {candidate.resolve() for candidate in candidates},
        key=lambda candidate: _candidate_score(candidate, image_path),
    )
    return unique[0] if unique else None


def _header_mapping(row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        canonical = _canonical_header(value)
        if canonical and canonical not in mapping:
            mapping[canonical] = index
    has_xyxy = {"x0", "x1"}.issubset(mapping)
    has_xywh = {"x", "width"}.issubset(mapping)
    return mapping if has_xyxy or has_xywh or "bbox" in mapping else {}


def _row_value(row, mapping: dict[str, int], key: str):
    index = mapping.get(key)
    return row[index] if index is not None and index < len(row) else None


def _bbox_numbers(value) -> tuple[float, float, float, float] | None:
    numbers = [float(item) for item in _NUMBER.findall(str(value or ""))]
    if len(numbers) < 4:
        return None
    x0, y0, third, fourth = numbers[:4]
    if os.environ.get("REAL_BOX_BBOX_FORMAT", "xyxy").strip().lower() == "xywh":
        return x0, y0, x0 + third, y0 + fourth
    return x0, y0, third, fourth


def _row_matches_image(row, mapping: dict[str, int], image_path: Path) -> bool:
    raw = _row_value(row, mapping, "image")
    if raw in (None, ""):
        return True
    value = Path(str(raw)).stem.lower()
    return value == image_path.stem.lower() or image_path.stem.lower() in value


def _parse_rows(rows: list[tuple], image_path: Path) -> list[SubwordBox]:
    header_index = -1
    mapping: dict[str, int] = {}
    for index, row in enumerate(rows[:20]):
        candidate = _header_mapping(row)
        if candidate:
            header_index, mapping = index, candidate
            break
    if not mapping:
        return []

    requested_line = _line_index(image_path)
    boxes: list[SubwordBox] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not _row_matches_image(row, mapping, image_path):
            continue
        if "line" in mapping and requested_line is not None:
            parsed_line = _as_float(_row_value(row, mapping, "line"))
            if parsed_line is not None and int(parsed_line) != requested_line:
                continue

        if "bbox" in mapping:
            parsed = _bbox_numbers(_row_value(row, mapping, "bbox"))
            if parsed is None:
                continue
            x0, y0, x1, y1 = parsed
        elif {"x0", "x1"}.issubset(mapping):
            x0 = _as_float(_row_value(row, mapping, "x0"))
            x1 = _as_float(_row_value(row, mapping, "x1"))
            y0 = _as_float(_row_value(row, mapping, "y0"))
            y1 = _as_float(_row_value(row, mapping, "y1"))
        else:
            x0 = _as_float(_row_value(row, mapping, "x"))
            y0 = _as_float(_row_value(row, mapping, "y"))
            width = _as_float(_row_value(row, mapping, "width"))
            height = _as_float(_row_value(row, mapping, "height"))
            x1 = None if x0 is None or width is None else x0 + width
            y1 = None if y0 is None or height is None else y0 + height

        if x0 is None or x1 is None:
            continue
        y0 = 0.0 if y0 is None else y0
        y1 = 1.0 if y1 is None else y1
        left, right = sorted((float(x0), float(x1)))
        top, bottom = sorted((float(y0), float(y1)))
        if right <= left:
            continue
        boxes.append(
            SubwordBox(
                text=str(_row_value(row, mapping, "text") or "").strip(),
                x0=left,
                y0=top,
                x1=right,
                y1=bottom,
                source_row=row_number,
            )
        )
    return boxes


def _sheet_matches_line(sheet_name: str, requested_line: int | None) -> bool:
    if requested_line is None:
        return True
    match = _LINE_NUMBER.search(str(sheet_name))
    return match is None or int(match.group(1)) == requested_line


def _read_workbook(workbook: Path, image_path: Path) -> BoxAnnotations:
    parsed: list[tuple[str, list[SubwordBox]]] = []
    try:
        if workbook.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            book = load_workbook(workbook, read_only=True, data_only=True)
            try:
                for sheet in book.worksheets:
                    if not _sheet_matches_line(sheet.title, _line_index(image_path)):
                        continue
                    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
                    boxes = _parse_rows(rows, image_path)
                    if boxes:
                        parsed.append((sheet.title, boxes))
            finally:
                book.close()
        else:
            import pandas as pd

            for sheet_name, frame in pd.read_excel(workbook, sheet_name=None).items():
                if not _sheet_matches_line(str(sheet_name), _line_index(image_path)):
                    continue
                rows = [tuple(frame.columns)] + list(frame.itertuples(index=False, name=None))
                boxes = _parse_rows(rows, image_path)
                if boxes:
                    parsed.append((str(sheet_name), boxes))
    except Exception as exc:
        return BoxAnnotations((), str(workbook), "", "parse_error", f"{type(exc).__name__}: {exc}")

    if not parsed:
        return BoxAnnotations((), str(workbook), "", "no_boxes", "No compatible subword box table found")
    sheet, boxes = max(parsed, key=lambda item: len(item[1]))
    return BoxAnnotations(tuple(boxes), str(workbook), sheet, "ok", "")


def _transcript_subwords(image_path: Path) -> list[str]:
    root = _side_root(image_path)
    candidates = (
        root / "text" / "final" / "original" / f"{image_path.stem}.txt",
        root / "text" / "final" / "tashkeel" / f"{image_path.stem}.txt",
        root / "text" / "raw" / f"{image_path.stem}.txt",
    )
    for candidate in candidates:
        if not candidate.is_file():
            continue
        text = candidate.read_text(encoding="utf-8").strip()
        try:
            from connected_subword_mode import connected_units

            return [unit.text for unit in connected_units(text) if unit.kind == "subword"]
        except Exception:
            return [token for token in text.split() if token]
    return []


def _reading_order(boxes) -> list[SubwordBox]:
    # Arabic logical order is right-to-left on the line image.
    return sorted(boxes, key=lambda box: (-box.center_x, box.y0, box.source_row))


def _fill_missing_text(boxes: list[SubwordBox], image_path: Path) -> list[SubwordBox]:
    if boxes and all(_normalise_text(box.text) for box in boxes):
        return boxes
    tokens = _transcript_subwords(image_path)
    if len(tokens) != len(boxes):
        return boxes
    return [
        SubwordBox(token, box.x0, box.y0, box.x1, box.y1, box.source_row)
        for token, box in zip(tokens, boxes)
    ]


def load_line_annotations(image_path, target_width: int, target_height: int) -> BoxAnnotations:
    image_path = Path(image_path)
    workbook = discover_workbook(str(image_path))
    if workbook is None:
        return BoxAnnotations((), "", "", "missing", "No Excel workbook found near the line image")
    annotations = _read_workbook(workbook, image_path)
    if not annotations.boxes:
        return annotations

    with Image.open(image_path) as opened:
        source_width, source_height = opened.size
    mode = os.environ.get("REAL_BOX_COORDINATE_SPACE", "original").strip().lower()
    maximum = max(max(box.x1, box.y1) for box in annotations.boxes)
    if mode == "auto":
        mode = "normalized" if maximum <= 1.5 else "original"
    if mode == "normalized":
        scale_x, scale_y = float(target_width), float(target_height)
    elif mode == "model":
        scale_x = scale_y = 1.0
    else:
        scale_x = float(target_width) / max(1.0, float(source_width))
        scale_y = float(target_height) / max(1.0, float(source_height))

    boxes = []
    for box in annotations.boxes:
        x0 = max(0.0, min(float(target_width), box.x0 * scale_x))
        x1 = max(0.0, min(float(target_width), box.x1 * scale_x))
        y0 = max(0.0, min(float(target_height), box.y0 * scale_y))
        y1 = max(0.0, min(float(target_height), box.y1 * scale_y))
        if y1 <= y0:
            y0, y1 = 0.0, float(target_height)
        if x1 > x0:
            boxes.append(SubwordBox(box.text, x0, y0, x1, y1, box.source_row))
    boxes = _fill_missing_text(_reading_order(boxes), image_path)
    status, error = annotations.status, annotations.error
    if boxes and not all(_normalise_text(box.text) for box in boxes):
        status = "missing_text"
        error = "Excel text is missing and transcript subword count does not match box count"
    return BoxAnnotations(tuple(boxes), annotations.workbook, annotations.sheet, status, error)


def _lcs_pairs(left: list[str], right: list[str]) -> list[tuple[int, int]]:
    n, m = len(left), len(right)
    dp = np.zeros((n + 1, m + 1), dtype=np.int32)
    for i in range(n - 1, -1, -1):
        for j in range(m - 1, -1, -1):
            dp[i, j] = (
                1 + dp[i + 1, j + 1]
                if left[i] and left[i] == right[j]
                else max(dp[i + 1, j], dp[i, j + 1])
            )
    pairs = []
    i = j = 0
    while i < n and j < m:
        if left[i] and left[i] == right[j] and dp[i, j] == 1 + dp[i + 1, j + 1]:
            pairs.append((i, j))
            i += 1
            j += 1
        elif dp[i + 1, j] >= dp[i, j + 1]:
            i += 1
        else:
            j += 1
    return pairs


def _ratio(numerator: float, denominator: float, empty: float = 0.0) -> float:
    return float(numerator / denominator) if denominator > 0 else float(empty)


def _binary_metrics(tp: int, fp: int, fn: int, tn: int) -> dict[str, float]:
    precision = _ratio(tp, tp + fp, 1.0 if tp + fn == 0 else 0.0)
    recall = _ratio(tp, tp + fn, 1.0)
    f1 = _ratio(2.0 * precision * recall, precision + recall, 0.0)
    return {
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "specificity": _ratio(tn, tn + fp, 1.0),
        "accuracy": _ratio(tp + tn, tp + fp + fn + tn, 0.0),
    }


def _box_coverage(box: SubwordBox, interval) -> float:
    if interval is None or box.width <= 0:
        return 0.0
    overlap = max(0.0, min(box.x1, interval[1]) - max(box.x0, interval[0]))
    return overlap / box.width


def _predicted_indices(boxes, interval) -> tuple[set[int], list[float]]:
    rule = os.environ.get("REAL_BOX_IN_MASK_RULE", "center").strip().lower()
    minimum = max(0.0, min(1.0, _number("REAL_BOX_MIN_COVERAGE", 0.50)))
    predicted, coverages = set(), []
    for index, box in enumerate(boxes):
        coverage = _box_coverage(box, interval)
        coverages.append(coverage)
        center_inside = bool(interval and interval[0] <= box.center_x <= interval[1])
        inside = center_inside if rule == "center" else coverage >= minimum
        if rule in {"center_or_coverage", "either"}:
            inside = center_inside or coverage >= minimum
        if inside:
            predicted.add(index)
    return predicted, coverages


def _union_interval(boxes, indices: set[int]):
    if not indices:
        return None
    return min(boxes[index].x0 for index in indices), max(boxes[index].x1 for index in indices)


def _interval_iou(left, right):
    if left is None or right is None:
        return None
    intersection = max(0.0, min(left[1], right[1]) - max(left[0], right[0]))
    union = max(left[1], right[1]) - min(left[0], right[0])
    return float(intersection / union) if union > 0 else 0.0


def _raster_metrics(boxes, gt_indices: set[int], interval, width: int, height: int):
    if not gt_indices:
        return {"precision": None, "recall": None, "f1": None, "iou": None}
    gt = np.zeros((height, width), dtype=bool)
    for index in gt_indices:
        box = boxes[index]
        x0, x1 = int(math.floor(box.x0)), int(math.ceil(box.x1))
        y0, y1 = int(math.floor(box.y0)), int(math.ceil(box.y1))
        gt[max(0, y0):min(height, y1), max(0, x0):min(width, x1)] = True
    pred = np.zeros_like(gt)
    if interval is not None:
        x0, x1 = int(math.floor(interval[0])), int(math.ceil(interval[1]))
        pred[:, max(0, x0):min(width, x1)] = True
    intersection = int(np.logical_and(gt, pred).sum())
    pred_area, gt_area = int(pred.sum()), int(gt.sum())
    union = int(np.logical_or(gt, pred).sum())
    precision = _ratio(intersection, pred_area, 1.0 if gt_area == 0 else 0.0)
    recall = _ratio(intersection, gt_area, 1.0)
    return {
        "precision": precision,
        "recall": recall,
        "f1": _ratio(2 * precision * recall, precision + recall, 0.0),
        "iou": _ratio(intersection, union, 1.0),
    }


def _predicted_interval(row: dict, prefix: str, width: int, use_flip: bool):
    try:
        start = int(row[f"{prefix}_path_start"])
        end = int(row[f"{prefix}_path_end"])
        windows = int(row[f"{prefix}_windows"])
    except (KeyError, TypeError, ValueError):
        return None
    if start < 0 or end < start or windows <= 0:
        return None
    return tuple(sorted(patch_range_to_pixels(start, end + 1, windows, width, use_flip)))


def _line_metrics(prefix: str, annotations: BoxAnnotations, gt_indices: set[int], interval, width: int, height: int):
    boxes = annotations.boxes
    predicted, coverages = _predicted_indices(boxes, interval)
    tp = len(predicted & gt_indices)
    fp = len(predicted - gt_indices)
    fn = len(gt_indices - predicted)
    tn = len(set(range(len(boxes))) - predicted - gt_indices)
    binary = _binary_metrics(tp, fp, fn, tn)
    gt_interval = _union_interval(boxes, gt_indices)
    interval_iou = _interval_iou(interval, gt_interval)
    raster = _raster_metrics(boxes, gt_indices, interval, width, height)
    return {
        f"{prefix}_box_annotation_path": annotations.workbook,
        f"{prefix}_box_annotation_sheet": annotations.sheet,
        f"{prefix}_box_annotation_status": annotations.status,
        f"{prefix}_box_annotation_error": annotations.error,
        f"{prefix}_box_count": len(boxes),
        f"{prefix}_shared_gt_boxes": len(gt_indices),
        f"{prefix}_predicted_mask_boxes": len(predicted),
        f"{prefix}_box_tp": tp,
        f"{prefix}_box_fp": fp,
        f"{prefix}_box_fn": fn,
        f"{prefix}_box_tn": tn,
        f"{prefix}_box_precision": binary["precision"],
        f"{prefix}_box_recall": binary["recall"],
        f"{prefix}_box_f1": binary["f1"],
        f"{prefix}_box_specificity": binary["specificity"],
        f"{prefix}_box_accuracy": binary["accuracy"],
        f"{prefix}_mean_box_mask_coverage": float(np.mean(coverages)) if coverages else None,
        f"{prefix}_shared_box_mask_coverage": (
            float(np.mean([coverages[index] for index in gt_indices])) if gt_indices else None
        ),
        f"{prefix}_box_interval_iou": interval_iou,
        f"{prefix}_box_pixel_precision": raster["precision"],
        f"{prefix}_box_pixel_recall": raster["recall"],
        f"{prefix}_box_pixel_f1": raster["f1"],
        f"{prefix}_box_pixel_iou": raster["iou"],
        f"{prefix}_pred_start_px": None if interval is None else float(interval[0]),
        f"{prefix}_pred_end_px": None if interval is None else float(interval[1]),
        f"{prefix}_gt_start_px": None if gt_interval is None else float(gt_interval[0]),
        f"{prefix}_gt_end_px": None if gt_interval is None else float(gt_interval[1]),
        f"{prefix}_region_iou": interval_iou,
        f"{prefix}_start_error_px": (
            None if interval is None or gt_interval is None else abs(interval[0] - gt_interval[0])
        ),
        f"{prefix}_end_error_px": (
            None if interval is None or gt_interval is None else abs(interval[1] - gt_interval[1])
        ),
    }


def metrics_from_evaluation_row(row: dict, pair, models) -> dict:
    """Add real box metrics to one completed sw_runner evaluation row."""
    if str(row.get("dataset_type", "")).lower() != "real" or not _flag("REAL_BOX_EVAL", True):
        return {"real_box_evaluated": False, "real_box_status": "disabled"}

    with Image.open(pair.image1) as opened:
        original_width1, original_height1 = opened.size
    with Image.open(pair.image2) as opened:
        original_width2, original_height2 = opened.size
    model_width = int(os.environ.get("LINE_WIDTH", "1024"))
    model_height = int(os.environ.get("LINE_HEIGHT", "128"))
    width1 = model_width if row.get("binarized") else original_width1
    height1 = model_height if row.get("binarized") else original_height1
    width2 = model_width if row.get("binarized") else original_width2
    height2 = model_height if row.get("binarized") else original_height2

    ann1 = load_line_annotations(pair.image1, width1, height1)
    ann2 = load_line_annotations(pair.image2, width2, height2)
    if _flag("REAL_REQUIRE_BOX_ANNOTATIONS", False) and (not ann1.boxes or not ann2.boxes):
        raise FileNotFoundError(f"Missing real box annotations: line1={ann1.error}; line2={ann2.error}")

    texts1 = [_normalise_text(box.text) for box in ann1.boxes]
    texts2 = [_normalise_text(box.text) for box in ann2.boxes]
    matches = _lcs_pairs(texts1, texts2) if ann1.boxes and ann2.boxes else []
    gt1 = {left for left, _ in matches}
    gt2 = {right for _, right in matches}
    use_flip = bool(models.image_model.use_flip)
    interval1 = _predicted_interval(row, "line1", width1, use_flip)
    interval2 = _predicted_interval(row, "line2", width2, use_flip)
    line1 = _line_metrics("line1", ann1, gt1, interval1, width1, height1)
    line2 = _line_metrics("line2", ann2, gt2, interval2, width2, height2)

    tp = int(line1["line1_box_tp"]) + int(line2["line2_box_tp"])
    fp = int(line1["line1_box_fp"]) + int(line2["line2_box_fp"])
    fn = int(line1["line1_box_fn"]) + int(line2["line2_box_fn"])
    tn = int(line1["line1_box_tn"]) + int(line2["line2_box_tn"])
    micro = _binary_metrics(tp, fp, fn, tn)
    interval_ious = [value for value in (line1["line1_region_iou"], line2["line2_region_iou"]) if value is not None]
    pixel_ious = [value for value in (line1["line1_box_pixel_iou"], line2["line2_box_pixel_iou"]) if value is not None]

    status = "ok"
    if not ann1.boxes or not ann2.boxes:
        status = "missing_annotations"
    elif ann1.status != "ok" or ann2.status != "ok":
        status = "annotation_warning"
    elif not matches:
        status = "no_shared_subword_boxes"

    return {
        "real_box_evaluated": bool(ann1.boxes and ann2.boxes),
        "real_box_status": status,
        "real_box_match_rule": os.environ.get("REAL_BOX_IN_MASK_RULE", "center"),
        "real_box_min_coverage": _number("REAL_BOX_MIN_COVERAGE", 0.50),
        "real_box_coordinate_space": os.environ.get("REAL_BOX_COORDINATE_SPACE", "original"),
        "shared_subword_matches": len(matches),
        "pair_box_tp": tp,
        "pair_box_fp": fp,
        "pair_box_fn": fn,
        "pair_box_tn": tn,
        "pair_box_precision": micro["precision"],
        "pair_box_recall": micro["recall"],
        "pair_box_f1": micro["f1"],
        "pair_box_specificity": micro["specificity"],
        "pair_box_accuracy": micro["accuracy"],
        "mean_box_interval_iou": float(np.mean(interval_ious)) if interval_ious else None,
        "mean_box_pixel_iou": float(np.mean(pixel_ious)) if pixel_ious else None,
        "mean_region_iou": float(np.mean(interval_ious)) if interval_ious else None,
        **line1,
        **line2,
    }


def fieldnames() -> list[str]:
    result = [
        "real_box_evaluated", "real_box_status", "real_box_match_rule",
        "real_box_min_coverage", "real_box_coordinate_space", "shared_subword_matches",
        "pair_box_tp", "pair_box_fp", "pair_box_fn", "pair_box_tn",
        "pair_box_precision", "pair_box_recall", "pair_box_f1",
        "pair_box_specificity", "pair_box_accuracy", "mean_box_interval_iou",
        "mean_box_pixel_iou",
    ]
    for prefix in ("line1", "line2"):
        result.extend([
            f"{prefix}_box_annotation_path", f"{prefix}_box_annotation_sheet",
            f"{prefix}_box_annotation_status", f"{prefix}_box_annotation_error",
            f"{prefix}_box_count", f"{prefix}_shared_gt_boxes",
            f"{prefix}_predicted_mask_boxes", f"{prefix}_box_tp", f"{prefix}_box_fp",
            f"{prefix}_box_fn", f"{prefix}_box_tn", f"{prefix}_box_precision",
            f"{prefix}_box_recall", f"{prefix}_box_f1", f"{prefix}_box_specificity",
            f"{prefix}_box_accuracy", f"{prefix}_mean_box_mask_coverage",
            f"{prefix}_shared_box_mask_coverage", f"{prefix}_box_interval_iou",
            f"{prefix}_box_pixel_precision", f"{prefix}_box_pixel_recall",
            f"{prefix}_box_pixel_f1", f"{prefix}_box_pixel_iou",
        ])
    return result


def aggregate(rows: list[dict]) -> dict:
    annotated = [row for row in rows if row.get("real_box_evaluated") in (True, "True", 1, "1")]
    tp = sum(int(float(row.get("pair_box_tp") or 0)) for row in annotated)
    fp = sum(int(float(row.get("pair_box_fp") or 0)) for row in annotated)
    fn = sum(int(float(row.get("pair_box_fn") or 0)) for row in annotated)
    tn = sum(int(float(row.get("pair_box_tn") or 0)) for row in annotated)
    micro = _binary_metrics(tp, fp, fn, tn)

    def values(key: str) -> list[float]:
        output = []
        for row in annotated:
            try:
                value = float(row.get(key))
            except (TypeError, ValueError):
                continue
            if math.isfinite(value):
                output.append(value)
        return output

    statuses: dict[str, int] = {}
    for row in rows:
        status = str(row.get("real_box_status") or "not_evaluated")
        statuses[status] = statuses.get(status, 0) + 1
    mean = lambda items: float(np.mean(items)) if items else None
    return {
        "real_box_samples": len(annotated),
        "real_box_missing_or_unusable_samples": len(rows) - len(annotated),
        "real_box_status_counts": statuses,
        "box_micro_tp": tp,
        "box_micro_fp": fp,
        "box_micro_fn": fn,
        "box_micro_tn": tn,
        "box_micro_precision": micro["precision"],
        "box_micro_recall": micro["recall"],
        "box_micro_f1": micro["f1"],
        "box_micro_specificity": micro["specificity"],
        "box_micro_accuracy": micro["accuracy"],
        "box_macro_precision": mean(values("pair_box_precision")),
        "box_macro_recall": mean(values("pair_box_recall")),
        "box_macro_f1": mean(values("pair_box_f1")),
        "mean_box_interval_iou": mean(values("mean_box_interval_iou")),
        "mean_box_pixel_iou": mean(values("mean_box_pixel_iou")),
        "mean_shared_subword_matches": mean(values("shared_subword_matches")),
    }
